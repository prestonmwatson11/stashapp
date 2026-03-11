"""
LNS Hours Writer — Stage 3
Reads lns_hours_approved.json and writes approved hours
back into the master Excel workbook.

Usage:
    python write_hours.py [--dry-run]
"""

import json
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

WORKBOOK_PATH = "US_Local_Needlepoint_Shops_Directory.xlsx"
APPROVED_JSON = "lns_hours_approved.json"
HEADER_ROW    = 3
DATA_START    = 4

# Light green fill to mark newly-added cells
NEW_FILL = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")


def main():
    dry_run = "--dry-run" in sys.argv

    if not Path(APPROVED_JSON).exists():
        print(f"ERROR: {APPROVED_JSON} not found. Run review_hours.py first.")
        return

    with open(APPROVED_JSON) as f:
        approved_list = json.load(f)

    approved = {
        r["shop_name"]: r["approved_hours"]
        for r in approved_list
        if r.get("approved_hours") and r.get("review_status") != "skipped"
    }
    print(f"Approved hours to write: {len(approved)}")

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["Master Dataset"]

    headers = [c.value for c in ws[HEADER_ROW]]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}  # 1-indexed

    name_col  = col_idx["Shop Name"]
    hours_col = col_idx["Hours of Operation"]

    updated = 0
    not_matched = list(approved.keys())

    for row in ws.iter_rows(min_row=DATA_START):
        name_cell  = row[name_col  - 1]
        hours_cell = row[hours_col - 1]

        shop_name = name_cell.value
        if not shop_name or str(shop_name).startswith("  "):
            continue

        if shop_name in approved:
            hours_val = approved[shop_name]
            if dry_run:
                print(f"  [DRY RUN] {shop_name} → {hours_val[:70]}")
            else:
                hours_cell.value = hours_val
                hours_cell.fill  = NEW_FILL
            updated += 1
            if shop_name in not_matched:
                not_matched.remove(shop_name)

    if not dry_run:
        wb.save(WORKBOOK_PATH)
        print(f"\n✓ Workbook updated — {updated} shops now have hours.")
        print(f"  Saved → {WORKBOOK_PATH}")
    else:
        print(f"\n[DRY RUN] Would update {updated} shops.")

    if not_matched:
        print(f"\n  ⚠ {len(not_matched)} approved names not matched in workbook:")
        for n in not_matched[:20]:
            print(f"    – {n}")


if __name__ == "__main__":
    main()
