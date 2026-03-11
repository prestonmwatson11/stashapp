"""
LNS Phone Scraper
Scrapes each shop's website for a phone number.
Outputs: lns_phone_results.json (raw results for review)

Usage:
    python3 scrape_phones.py
"""

import json
import re
import time
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from typing import Optional, List
import openpyxl
from pathlib import Path

# ── Config ───────────────────────────────────────────────────────────────────
WORKBOOK_PATH = "US_Local_Needlepoint_Shops_Directory.xlsx"
OUTPUT_JSON   = "lns_phone_results.json"
APPROVED_JSON = "lns_phone_approved.json"
HEADER_ROW    = 3
DATA_START    = 4
REQUEST_DELAY = 1.5
TIMEOUT       = 10

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# Sub-pages most likely to have contact info
CONTACT_PAGE_HINTS = [
    "contact", "contact-us", "about", "location", "find-us",
    "visit", "store", "info", "directions", "reach-us",
]

# Phone number regex — matches most US formats:
#   (512) 555-1234  /  512.555.1234  /  512-555-1234  /  +1 512 555 1234
PHONE_RE = re.compile(
    r'(?:\+1[\s\-.]?)?\(?\d{3}\)?[\s\-.]?\d{3}[\s\-.]?\d{4}'
)

# Patterns to avoid — these look like phones but aren't
FALSE_POSITIVES = re.compile(
    r'(?:zip|postal|code|fax|ext|#\d{4}|\d{5}-\d{4})',
    re.IGNORECASE,
)

# ── Helpers ──────────────────────────────────────────────────────────────────

def normalise_url(raw: str) -> str:
    raw = raw.strip()
    if not raw.startswith("http"):
        raw = "https://" + raw
    return raw


def fetch(url: str, session: requests.Session) -> Optional[str]:
    try:
        r = session.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        if r.status_code == 200 and "text/html" in r.headers.get("content-type", ""):
            return r.text
    except Exception:
        pass
    return None


def find_contact_links(soup: BeautifulSoup, base_url: str) -> List[str]:
    candidates = []
    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        text = a.get_text(strip=True).lower()
        if any(hint in href or hint in text for hint in CONTACT_PAGE_HINTS):
            full = urljoin(base_url, a["href"])
            if urlparse(full).netloc == urlparse(base_url).netloc:
                candidates.append(full)
    return list(dict.fromkeys(candidates))[:4]


def normalise_phone(raw: str) -> str:
    """Standardise to (XXX) XXX-XXXX format."""
    digits = re.sub(r'\D', '', raw)
    if digits.startswith('1') and len(digits) == 11:
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return raw.strip()


def extract_phones(html: str) -> List[str]:
    """Return deduplicated list of phone numbers found in the page."""
    soup = BeautifulSoup(html, "lxml")

    # Remove script/style noise
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()

    found = []

    # 1. schema.org telephone markup
    for el in soup.find_all(attrs={"itemprop": re.compile("telephone", re.I)}):
        t = el.get("content") or el.get_text(strip=True)
        if t:
            found.append(normalise_phone(t.strip()))

    # 2. <a href="tel:..."> links
    for a in soup.find_all("a", href=re.compile(r'^tel:', re.I)):
        number = a["href"].replace("tel:", "").strip()
        if number:
            found.append(normalise_phone(number))

    if found:
        return list(dict.fromkeys(found))

    # 3. Regex sweep over page text
    text = soup.get_text(" ", strip=True)
    for m in PHONE_RE.finditer(text):
        snippet = m.group(0)
        context = text[max(0, m.start()-30):m.end()+30]
        # Skip if surrounded by false-positive context
        if FALSE_POSITIVES.search(context):
            continue
        found.append(normalise_phone(snippet))

    return list(dict.fromkeys(found))[:3]  # return up to 3 candidates


def scrape_shop(name: str, city: str, state: str, url: str,
                session: requests.Session) -> dict:
    result = {
        "shop_name": name,
        "city": city,
        "state": state,
        "url": url,
        "status": "pending",
        "phones_found": [],
        "best_guess": None,
        "source_page": None,
    }

    norm_url = normalise_url(url)

    # Step 1: homepage
    html = fetch(norm_url, session)
    if not html:
        result["status"] = "unreachable"
        return result

    phones = extract_phones(html)
    if phones:
        result["status"] = "found_homepage"
        result["phones_found"] = phones
        result["best_guess"] = phones[0]
        result["source_page"] = norm_url
        return result

    # Step 2: contact/about sub-pages
    soup = BeautifulSoup(html, "lxml")
    sub_links = find_contact_links(soup, norm_url)

    for link in sub_links:
        time.sleep(0.5)
        sub_html = fetch(link, session)
        if sub_html:
            phones = extract_phones(sub_html)
            if phones:
                result["status"] = "found_subpage"
                result["phones_found"] = phones
                result["best_guess"] = phones[0]
                result["source_page"] = link
                return result

    result["status"] = "not_found"
    return result


# ── Workbook loader ──────────────────────────────────────────────────────────

def load_shops() -> List[dict]:
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["Master Dataset"]
    headers = [c.value for c in ws[HEADER_ROW]]
    col = {h: i for i, h in enumerate(headers) if h}

    shops = []
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        name  = row[col["Shop Name"]]
        url   = row[col["Website URL"]]
        phone = row[col["Phone Number"]]
        city  = row[col["City"]]
        state = row[col["State"]]

        if not name or str(name).startswith("  "):
            continue
        if phone:
            continue  # already have a number
        if not url:
            shops.append({
                "shop_name": name, "city": city, "state": state,
                "url": None, "status": "no_url",
                "phones_found": [], "best_guess": None, "source_page": None,
            })
            continue
        shops.append({"shop_name": name, "city": city, "state": state, "url": str(url)})
    return shops


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    shops = load_shops()
    print(f"Shops to scrape: {len([s for s in shops if s.get('url')])} (missing phone, have URL)")

    # Resume support
    existing = {}
    if Path(OUTPUT_JSON).exists():
        with open(OUTPUT_JSON) as f:
            for item in json.load(f):
                existing[item["shop_name"]] = item
        print(f"Resuming — {len(existing)} already processed")

    session  = requests.Session()
    results  = list(existing.values())
    done     = set(existing.keys())

    to_scrape = [s for s in shops if s["shop_name"] not in done and s.get("url")]
    no_url    = [s for s in shops if not s.get("url") and s["shop_name"] not in done]
    results.extend(no_url)

    print(f"Scraping {len(to_scrape)} shops...\n")

    for i, shop in enumerate(to_scrape, 1):
        print(f"[{i}/{len(to_scrape)}] {shop['shop_name']} ({shop['city']}, {shop['state']})")
        result = scrape_shop(
            shop["shop_name"], shop["city"], shop["state"], shop["url"], session
        )
        results.append(result)

        with open(OUTPUT_JSON, "w") as f:
            json.dump(results, f, indent=2)

        icon = {"found_homepage": "✓", "found_subpage": "✓✓",
                "not_found": "–", "unreachable": "✗"}.get(result["status"], "?")
        print(f"   {icon} {result['status']}"
              + (f" → {result['best_guess']}" if result["best_guess"] else ""))

        time.sleep(REQUEST_DELAY)

    # Summary
    statuses = {}
    for r in results:
        statuses[r["status"]] = statuses.get(r["status"], 0) + 1
    print("\n── Summary ──")
    for k, v in sorted(statuses.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")
    print(f"\nResults saved → {OUTPUT_JSON}")
    print("Next step: python3 write_phones.py --auto-approve")


if __name__ == "__main__":
    main()
