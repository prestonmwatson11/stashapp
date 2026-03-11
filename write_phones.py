"""
LNS Phone Writer
Reads lns_phone_results.json, auto-approves clean phone numbers,
and writes them into the master Excel workbook.

Usage:
    python3 write_phones.py             # interactive review
    python3 write_phones.py --auto-approve   # approve all, skip review
    python3 write_phones.py --dry-run        # preview without saving
"""

import json
import re
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

WORKBOOK_PATH = "US_Local_Needlepoint_Shops_Directory.xlsx"
RESULTS_JSON  = "lns_phone_results.json"
APPROVED_JSON = "lns_phone_approved.json"
HEADER_ROW    = 3
DATA_START    = 4

NEW_FILL = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")

CLEAN_PHONE = re.compile(r'^\(\d{3}\) \d{3}-\d{4}$')


def looks_clean(phone: str) -> bool:
    return bool(CLEAN_PHONE.match(phone or ""))


def review_and_approve(results: list) -> dict:
    auto = "--auto-approve" in sys.argv
    approved = {}

    found = [r for r in results if r["status"] in ("found_homepage", "found_subpage")]
    print(f"\nFound phone numbers: {len(found)}")

    for r in found:
        phone = r.get("best_guess", "")
        if auto or looks_clean(phone):
            approved[r["shop_name"]] = phone
        else:
            print(f"\n  {r['shop_name']} ({r['city']}, {r['state']})")
            print(f"  Raw result: {phone}")
            print(f"  All found:  {r.get('phones_found', [])}")
            choice = input("  Accept? [Enter=yes / type correction / s=skip]: ").strip()
            if choice.lower() == "s":
                continue
            elif choice:
                approved[r["shop_name"]] = choice
            else:
                approved[r["shop_name"]] = phone

    with open(APPROVED_JSON, "w") as f:
        json.dump([{"shop_name": k, "phone": v} for k, v in approved.items()], f, indent=2)

    return approved


def write_to_workbook(approved: dict):
    dry_run = "--dry-run" in sys.argv

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["Master Dataset"]

    headers  = [c.value for c in ws[HEADER_ROW]]
    col_idx  = {h: i + 1 for i, h in enumerate(headers) if h}
    name_col  = col_idx["Shop Name"]
    phone_col = col_idx["Phone Number"]

    updated = 0
    for row in ws.iter_rows(min_row=DATA_START):
        name = row[name_col - 1].value
        if not name or str(name).startswith("  "):
            continue
        if name in approved:
            if dry_run:
                print(f"  [DRY RUN] {name} → {approved[name]}")
            else:
                cell = row[phone_col - 1]
                cell.value = approved[name]
                cell.fill  = NEW_FILL
            updated += 1

    if not dry_run:
        wb.save(WORKBOOK_PATH)
        print(f"\n✓ Workbook updated — {updated} shops now have phone numbers.")
    else:
        print(f"\n[DRY RUN] Would update {updated} shops.")


def main():
    if not Path(RESULTS_JSON).exists():
        print(f"ERROR: {RESULTS_JSON} not found. Run scrape_phones.py first.")
        return

    with open(RESULTS_JSON) as f:
        results = json.load(f)

    statuses = {}
    for r in results:
        statuses[r["status"]] = statuses.get(r["status"], 0) + 1
    print("── Scrape Results ──")
    for k, v in sorted(statuses.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")

    approved = review_and_approve(results)
    write_to_workbook(approved)
    print(f"Approved JSON saved → {APPROVED_JSON}")


if __name__ == "__main__":
    main()
