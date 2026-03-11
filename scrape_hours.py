"""
LNS Hours Scraper — Stage 1
Scrapes each shop's website looking for hours of operation.
Outputs: lns_hours_results.json (raw results for review)
"""

import json
import re
import time
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from typing import Optional, List
import openpyxl
from pathlib import Path

# ── Config ──────────────────────────────────────────────────────────────────
WORKBOOK_PATH = "US_Local_Needlepoint_Shops_Directory.xlsx"
OUTPUT_JSON   = "lns_hours_results.json"
HEADER_ROW    = 3          # 1-indexed row where column headers live
DATA_START    = 4          # first data row
REQUEST_DELAY = 1.5        # seconds between requests (be polite)
TIMEOUT       = 10         # seconds per request

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# Pages most likely to contain hours info
HOURS_PAGE_HINTS = [
    "contact", "contact-us", "hours", "visit", "about", "location",
    "find-us", "store", "info", "directions",
]

# Regex patterns to extract hours text
HOURS_PATTERNS = [
    # Mon–Fri 10am–5pm  /  Monday through Friday 10:00–17:00
    re.compile(
        r'(?:mon(?:day)?|tue(?:sday)?|wed(?:nesday)?|thu(?:rsday)?|'
        r'fri(?:day)?|sat(?:urday)?|sun(?:day)?)'
        r'[\s\-–—through to]+.*?'
        r'(?:\d{1,2}(?::\d{2})?\s*(?:am|pm)|closed)',
        re.IGNORECASE | re.DOTALL,
    ),
    # "Hours: Mon-Fri 10-5"
    re.compile(
        r'hours?\s*[:\-]?\s*.{5,80}',
        re.IGNORECASE,
    ),
    # "Open Monday"  /  "We are open"
    re.compile(
        r'(?:open|closed)\s+(?:mon|tue|wed|thu|fri|sat|sun).{0,60}',
        re.IGNORECASE,
    ),
    # "10am – 5pm" standalone time range
    re.compile(
        r'\d{1,2}(?::\d{2})?\s*(?:am|pm)\s*[–\-—to]+\s*\d{1,2}(?::\d{2})?\s*(?:am|pm)',
        re.IGNORECASE,
    ),
]

# ── Helpers ──────────────────────────────────────────────────────────────────

def normalise_url(raw: str) -> str:
    """Ensure URL has a scheme."""
    raw = raw.strip()
    if not raw.startswith("http"):
        raw = "https://" + raw
    return raw


def fetch(url: str, session: requests.Session) -> Optional[str]:
    """Fetch a URL; return text or None on failure."""
    try:
        r = session.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        if r.status_code == 200 and "text/html" in r.headers.get("content-type", ""):
            return r.text
    except Exception:
        pass
    return None


def find_hours_links(soup: BeautifulSoup, base_url: str) -> List[str]:
    """Return candidate sub-page URLs that might contain hours."""
    candidates = []
    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        text = a.get_text(strip=True).lower()
        if any(hint in href or hint in text for hint in HOURS_PAGE_HINTS):
            full = urljoin(base_url, a["href"])
            if urlparse(full).netloc == urlparse(base_url).netloc:
                candidates.append(full)
    return list(dict.fromkeys(candidates))[:4]  # dedupe, max 4


def extract_hours_text(html: str) -> List[str]:
    """Return a list of candidate hours strings found in the page."""
    soup = BeautifulSoup(html, "lxml")

    # Remove nav / footer / script noise
    for tag in soup(["script", "style", "noscript", "header", "nav"]):
        tag.decompose()

    # 1. Look for schema.org openingHours markup
    schema_hits = []
    for el in soup.find_all(attrs={"itemprop": re.compile("opening", re.I)}):
        t = el.get("content") or el.get_text(strip=True)
        if t:
            schema_hits.append(t.strip())
    if schema_hits:
        return schema_hits

    # 2. Look for elements with class/id containing 'hour' or 'schedule'
    keyword_hits = []
    for el in soup.find_all(
        True,
        attrs={
            "class": re.compile(r"hour|schedule|open|location|contact", re.I),
            # id is checked separately
        },
    ):
        text = el.get_text(" ", strip=True)
        if 10 < len(text) < 400:
            keyword_hits.append(text)
    for el in soup.find_all(
        True,
        id=re.compile(r"hour|schedule|open|location|contact", re.I),
    ):
        text = el.get_text(" ", strip=True)
        if 10 < len(text) < 400:
            keyword_hits.append(text)
    if keyword_hits:
        return keyword_hits[:3]

    # 3. Regex sweep over full page text
    page_text = soup.get_text(" ", strip=True)
    regex_hits = []
    for pattern in HOURS_PATTERNS:
        for m in pattern.finditer(page_text):
            snippet = m.group(0).strip()
            if len(snippet) > 8:
                # Grab a little context either side
                start = max(0, m.start() - 20)
                end   = min(len(page_text), m.end() + 60)
                context = page_text[start:end].strip()
                regex_hits.append(context)
    return list(dict.fromkeys(regex_hits))[:5]


def scrape_shop(name: str, city: str, state: str, url: str,
                session: requests.Session) -> dict:
    """Full scrape pipeline for one shop. Returns result dict."""
    result = {
        "shop_name": name,
        "city": city,
        "state": state,
        "url": url,
        "status": "pending",
        "hours_found": [],
        "best_guess": None,
        "source_page": None,
    }

    norm_url = normalise_url(url)

    # Step 1: fetch homepage
    html = fetch(norm_url, session)
    if not html:
        result["status"] = "unreachable"
        return result

    hours = extract_hours_text(html)
    if hours:
        result["status"] = "found_homepage"
        result["hours_found"] = hours
        result["best_guess"] = hours[0]
        result["source_page"] = norm_url
        return result

    # Step 2: try sub-pages (contact, about, etc.)
    soup = BeautifulSoup(html, "lxml")
    sub_links = find_hours_links(soup, norm_url)

    for link in sub_links:
        time.sleep(0.5)
        sub_html = fetch(link, session)
        if sub_html:
            hours = extract_hours_text(sub_html)
            if hours:
                result["status"] = "found_subpage"
                result["hours_found"] = hours
                result["best_guess"] = hours[0]
                result["source_page"] = link
                return result

    result["status"] = "not_found"
    return result


# ── Main ─────────────────────────────────────────────────────────────────────

def load_shops() -> list[dict]:
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["Master Dataset"]
    headers = [c.value for c in ws[HEADER_ROW]]
    col = {h: i for i, h in enumerate(headers) if h}

    shops = []
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        name  = row[col["Shop Name"]]
        url   = row[col["Website URL"]]
        hours = row[col["Hours of Operation"]]
        city  = row[col["City"]]
        state = row[col["State"]]

        # Skip section-header rows and shops already with hours
        if not name or str(name).startswith("  "):
            continue
        if hours:
            continue
        if not url:
            shops.append({
                "shop_name": name,
                "city": city,
                "state": state,
                "url": None,
                "status": "no_url",
                "hours_found": [],
                "best_guess": None,
                "source_page": None,
            })
            continue

        shops.append({
            "shop_name": name,
            "city": city,
            "state": state,
            "url": str(url),
        })
    return shops


def main():
    Path(".").mkdir(exist_ok=True)
    shops = load_shops()
    print(f"Shops to scrape: {len(shops)} (missing hours, have URL)")

    # Resume from existing results
    existing = {}
    if Path(OUTPUT_JSON).exists():
        with open(OUTPUT_JSON) as f:
            for item in json.load(f):
                existing[item["shop_name"]] = item
        print(f"Resuming — {len(existing)} already processed")

    session = requests.Session()
    results = list(existing.values())
    done_names = set(existing.keys())

    to_scrape = [s for s in shops if s["shop_name"] not in done_names and s.get("url")]
    no_url    = [s for s in shops if not s.get("url") and s["shop_name"] not in done_names]

    # Add no-URL shops immediately
    for s in no_url:
        results.append(s)

    print(f"Scraping {len(to_scrape)} shops...")
    for i, shop in enumerate(to_scrape, 1):
        print(f"[{i}/{len(to_scrape)}] {shop['shop_name']} ({shop['city']}, {shop['state']}) — {shop['url']}")
        result = scrape_shop(
            shop["shop_name"], shop["city"], shop["state"], shop["url"], session
        )
        results.append(result)

        # Save after every shop (safe resume)
        with open(OUTPUT_JSON, "w") as f:
            json.dump(results, f, indent=2)

        status_icon = {"found_homepage": "✓", "found_subpage": "✓✓",
                       "not_found": "–", "unreachable": "✗"}.get(result["status"], "?")
        print(f"   {status_icon} {result['status']}"
              + (f" → {result['best_guess'][:80]}" if result["best_guess"] else ""))

        time.sleep(REQUEST_DELAY)

    # Summary
    statuses = {}
    for r in results:
        statuses[r["status"]] = statuses.get(r["status"], 0) + 1
    print("\n── Summary ──")
    for k, v in sorted(statuses.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")
    print(f"\nResults saved → {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
