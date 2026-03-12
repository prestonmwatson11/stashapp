"""
LNS Combined Writer
Reads lns_combined_approved.json (exported from the browser editor)
and writes hours, phone, Instagram, and website URL into the master
Excel workbook. Skips shops marked as deleted.

Usage:
    python3 write_combined.py              # write only empty cells
    python3 write_combined.py --overwrite  # overwrite existing values too
    python3 write_combined.py --dry-run    # preview without saving
    python3 write_combined.py --deleted    # list shops marked as deleted
"""

import json
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

WORKBOOK_PATH = "US_Local_Needlepoint_Shops_Directory.xlsx"
COMBINED_JSON = "lns_combined_approved.json"
HEADER_ROW    = 3
DATA_START    = 4

NEW_FILL     = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
DELETED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")


def main():
    dry_run   = "--dry-run"   in sys.argv
    overwrite = "--overwrite" in sys.argv
    show_del  = "--deleted"   in sys.argv

    if not Path(COMBINED_JSON).exists():
        print(f"ERROR: {COMBINED_JSON} not found. Export it from lns_editor.html first.")
        return

    with open(COMBINED_JSON) as f:
        data = json.load(f)

    deleted  = {r["shop_name"] for r in data if r.get("deleted")}
    approved = {r["shop_name"]: r for r in data if not r.get("deleted")}

    print(f"Active shops  : {len(approved)}")
    print(f"Deleted shops : {len(deleted)}")

    if show_del:
        print("\nShops marked as deleted:")
        for name in sorted(deleted):
            print(f"  - {name}")
        return

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["Master Dataset"]

    headers = [c.value for c in ws[HEADER_ROW]]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}
    name_col  = col_idx["Shop Name"]
    url_col   = col_idx["Website URL"]
    hours_col = col_idx["Hours of Operation"]
    phone_col = col_idx["Phone Number"]
    ig_col    = col_idx["Instagram Handle"]

    counts = {"URL": 0, "HOURS": 0, "PHONE": 0, "INSTA": 0, "DELETED": 0}

    for row in ws.iter_rows(min_row=DATA_START):
        name = row[name_col - 1].value
        if not name or str(name).startswith("  "):
            continue

        if name in deleted:
            if not dry_run:
                for cell in row:
                    if cell.value is not None:
                        cell.fill = DELETED_FILL
            else:
                print(f"  [DELETED] {name}")
            counts["DELETED"] += 1
            continue

        if name not in approved:
            continue

        entry = approved[name]

        def write(col, value, key):
            cell = row[col - 1]
            if value and (overwrite or not cell.value):
                if dry_run:
                    print(f"  [{key}] {name} -> {str(value)[:70]}")
                else:
                    cell.value = value
                    cell.fill  = NEW_FILL
                counts[key] += 1

        write(url_col,   entry.get("url"),         "URL")
        write(hours_col, entry.get("hours_value"), "HOURS")
        write(phone_col, entry.get("phone_value"), "PHONE")
        write(ig_col,    entry.get("ig_value"),    "INSTA")

    if not dry_run:
        wb.save(WORKBOOK_PATH)
        print(f"\nWorkbook updated:")
        print(f"  Website written  : {counts['URL']}")
        print(f"  Hours written    : {counts['HOURS']}")
        print(f"  Phones written   : {counts['PHONE']}")
        print(f"  Instagram written: {counts['INSTA']}")
        print(f"  Deleted flagged  : {counts['DELETED']}")
        print(f"\n  Saved -> {WORKBOOK_PATH}")
    else:
        print(f"\n[DRY RUN] Would write: {counts['URL']} URLs, {counts['HOURS']} hours, "
              f"{counts['PHONE']} phones, {counts['INSTA']} IG handles, "
              f"flag {counts['DELETED']} deleted.")


if __name__ == "__main__":
    main()
